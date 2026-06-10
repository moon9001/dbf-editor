#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
DBF 编辑器服务端 v2
支持 DBF 文件的查看、编辑、导入导出、转 Excel
固定端口 5000，支持局域网访问
"""

import os
import io
import json
import stat
import shutil
import time
from pathlib import Path

import dbf
from flask import Flask, request, jsonify, send_file

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB


@app.after_request
def add_no_cache(response):
    """禁止缓存（确保前端总是加载最新代码）"""
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / 'uploads'
TEMPLATES_DIR = BASE_DIR / 'templates'
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

PORT = 5000  # 固定端口


# ========== 工具函数 ==========

def _safe_open_read(filepath):
    """安全打开 DBF 只读"""
    table = dbf.Table(str(filepath), codepage='cp936')
    table.open(dbf.READ_ONLY)
    return table


def _safe_open_write(filepath):
    """安全打开 DBF 读写（可能需要在副本上操作）"""
    path = str(filepath)
    # 确保文件可写
    if os.path.exists(path):
        try:
            os.chmod(path, stat.S_IWRITE | stat.S_IREAD)
        except Exception:
            pass
    table = dbf.Table(path, codepage='cp936')
    table.open(dbf.READ_WRITE)
    return table


def _copy_file(src, dst):
    """复制文件并确保目标可写"""
    with open(src, 'rb') as fsrc:
        data = fsrc.read()
    with open(dst, 'wb') as fdst:
        fdst.write(data)
    try:
        os.chmod(dst, stat.S_IWRITE | stat.S_IREAD)
    except Exception:
        pass


def _field_type_char(code):
    """将 field_info 返回的类型码转为字符"""
    if isinstance(code, int):
        return chr(code)
    return str(code)


def get_dbf_info(filepath):
    """获取 DBF 文件的结构信息"""
    try:
        table = _safe_open_read(filepath)
        fields = []
        for fname in table.field_names:
            info = table.field_info(fname)
            fields.append({
                'name': fname,
                'type': _field_type_char(info[0]),
                'length': info[1],
                'decimal': info[2]
            })
        count = len(table)
        table.close()
        return {
            'filename': os.path.basename(filepath),
            'filepath': str(filepath),
            'fields': fields,
            'field_count': len(fields),
            'record_count': count,
            'encoding': 'gbk',
            'file_size': os.path.getsize(filepath)
        }
    except Exception as e:
        return {'error': str(e)}


def get_dbf_records(filepath, page=1, page_size=50, search='', search_field='', filters=None):
    """分页读取 DBF 记录
    filters: dict，额外的精确匹配筛选，如 {'DWDM': '80106'}
    """
    try:
        table = _safe_open_read(filepath)
        field_names = list(table.field_names)

        all_records = []
        for rec in table:
            if dbf.is_deleted(rec):
                continue
            row = {}
            for fld in field_names:
                try:
                    val = getattr(rec, fld, '')
                    if isinstance(val, bytes):
                        val = val.decode('gbk', errors='replace')
                    elif val is None:
                        val = ''
                    else:
                        val = str(val)
                    row[fld] = val.strip()
                except Exception:
                    row[fld] = ''
            all_records.append(row)

        total = len(all_records)

        # 额外精确筛选（如按 DWDM 筛选）
        if filters:
            for fld, val in filters.items():
                if val and val != '__all__':
                    all_records = [r for r in all_records if (r.get(fld, '') or '') == val]

        # 搜索过滤
        if search:
            if search_field and search_field != '__all__':
                all_records = [r for r in all_records if search.lower() in (r.get(search_field, '') or '').lower()]
            else:
                all_records = [r for r in all_records if any(
                    search.lower() in (v or '') for v in r.values()
                )]

        filtered_total = len(all_records)

        # 分页
        start = (page - 1) * page_size
        end = start + page_size
        page_records = all_records[start:end]

        table.close()
        return {
            'records': page_records,
            'total': total,
            'filtered_total': filtered_total,
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (filtered_total + page_size - 1) // page_size),
            'field_names': field_names
        }
    except Exception as e:
        return {'error': str(e)}


def update_dbf_record(filepath, index, updates):
    """更新 DBF 中的一条记录（直接原地更新）"""
    table = None
    try:
        table = _safe_open_write(filepath)
        if index >= len(table):
            table.close()
            return {'error': '行索引超出范围'}

        with table[index] as rec:
            for key, val in updates.items():
                if key in table.field_names:
                    try:
                        setattr(rec, key, str(val) if val is not None else '')
                    except Exception:
                        pass
        table.close()
        return {'success': True}
    except Exception as e:
        if table:
            try: table.close()
            except: pass
        return {'error': str(e)}


def add_dbf_record(filepath, record_data):
    """向 DBF 添加一条新记录"""
    table = None
    try:
        table = _safe_open_write(filepath)
        row_data = []
        for fld in table.field_names:
            val = record_data.get(fld, '')
            row_data.append(str(val) if val is not None else '')
        table.append(tuple(row_data))
        new_count = len(table)
        table.close()
        return {'success': True, 'index': new_count - 1}
    except Exception as e:
        if table:
            try: table.close()
            except: pass
        return {'error': str(e)}


def delete_dbf_record(filepath, index):
    """删除 DBF 中的一条记录"""
    table = None
    try:
        table = _safe_open_write(filepath)
        if index >= len(table):
            table.close()
            return {'error': '行索引超出范围'}

        dbf.delete(table[index])
        table.pack()
        table.close()
        return {'success': True}
    except Exception as e:
        if table:
            try: table.close()
            except: pass
        return {'error': str(e)}


def create_dbf_file(filepath, fields_info):
    """创建一个新的 DBF 文件"""
    try:
        spec_parts = []
        for f in fields_info:
            name = f.get('name', 'FIELD')
            ftype = f.get('type', 'C')
            length = int(f.get('length', 50))
            decimal = int(f.get('decimal', 0))
            if ftype == 'D':
                spec_parts.append(f'{name} D')
            elif ftype == 'L':
                spec_parts.append(f'{name} L')
            elif ftype == 'M':
                spec_parts.append(f'{name} M')
            elif ftype == 'N' and decimal > 0:
                spec_parts.append(f'{name} N({length},{decimal})')
            elif ftype == 'N':
                spec_parts.append(f'{name} N({length},0)')
            else:
                spec_parts.append(f'{name} C({length})')

        spec = '; '.join(spec_parts)
        table = dbf.Table(str(filepath), spec, codepage='cp936')
        table.open(dbf.READ_WRITE)
        table.close()
        return {'success': True}
    except Exception as e:
        return {'error': str(e)}


# ========== API 路由 ==========

@app.route('/api/distinct')
def get_distinct_values():
    """获取某字段的所有不重复值（用于筛选下拉）"""
    filepath = request.args.get('file', '')
    field = request.args.get('field', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})
    if not field:
        return jsonify({'error': '未指定字段'})
    try:
        table = _safe_open_read(filepath)
        field_upper = field.upper()
        if field_upper not in [f.upper() for f in table.field_names]:
            table.close()
            return jsonify({'error': f'字段 {field} 不存在'})
        # 找到实际字段名（区分大小写）
        actual_field = next(f for f in table.field_names if f.upper() == field_upper)
        vals = set()
        for rec in table:
            if dbf.is_deleted(rec):
                continue
            try:
                v = getattr(rec, actual_field.lower(), '')
                if isinstance(v, bytes):
                    v = v.decode('gbk', errors='replace')
                v = str(v).strip() if v is not None else ''
                if v:
                    vals.add(v)
            except Exception:
                pass
        table.close()
        sorted_vals = sorted(vals, key=lambda x: x)
        return jsonify({'field': actual_field, 'values': sorted_vals, 'count': len(sorted_vals)})
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/files')
def list_files():
    """列出所有可用的 DBF 文件"""
    files = []
    # 扫描 uploads 目录
    for f in sorted(UPLOAD_DIR.glob('*.dbf')):
        files.append({
            'name': f.name,
            'path': str(f.absolute()),
            'size': f.stat().st_size,
            'source': 'uploads'
        })
    # 扫描父目录下的 DBF 文件（默认数据文件）
    for f in sorted(BASE_DIR.parent.glob('*.dbf')):
        files.append({
            'name': f.name,
            'path': str(f.absolute()),
            'size': f.stat().st_size,
            'source': 'default'
        })
    return jsonify({'files': files})


@app.route('/api/info')
def get_info():
    """获取 DBF 文件结构"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})
    return jsonify(get_dbf_info(filepath))


@app.route('/api/data')
def get_data():
    """获取分页数据"""
    filepath = request.args.get('file', '')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    search = request.args.get('search', '')
    search_field = request.args.get('search_field', '')

    # 解析额外筛选条件，格式：filter_DWDM=80106
    filters = {}
    for key, val in request.args.items():
        if key.startswith('filter_') and val and val != '__all__':
            field_name = key[7:]  # 去掉 'filter_' 前缀
            filters[field_name] = val

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    return jsonify(get_dbf_records(filepath, page, page_size, search, search_field, filters))


@app.route('/api/update', methods=['POST'])
def update_record():
    """更新一条记录"""
    data = request.json
    filepath = data.get('file', '')
    index = int(data.get('index', -1))
    updates = data.get('updates', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    return jsonify(update_dbf_record(filepath, index, updates))


@app.route('/api/add', methods=['POST'])
def add_record():
    """添加一条记录"""
    data = request.json
    filepath = data.get('file', '')
    record = data.get('record', {})

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    return jsonify(add_dbf_record(filepath, record))


@app.route('/api/delete', methods=['POST'])
def delete_record():
    """删除一条记录"""
    data = request.json
    filepath = data.get('file', '')
    index = int(data.get('index', -1))

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    return jsonify(delete_dbf_record(filepath, index))


@app.route('/api/batch_delete', methods=['POST'])
def batch_delete():
    """批量删除"""
    data = request.json
    filepath = data.get('file', '')
    indices = sorted(data.get('indices', []), reverse=True)

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    errors = []
    for idx in indices:
        result = delete_dbf_record(filepath, idx)
        if 'error' in result:
            errors.append(f'索引 {idx}: {result["error"]}')

    if errors:
        return jsonify({'error': '; '.join(errors)})
    return jsonify({'success': True})


@app.route('/api/export/excel')
def export_excel():
    """导出为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})

    try:
        result = get_dbf_records(filepath, 1, 999999)
        if 'error' in result:
            return jsonify(result)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '数据'

        # 样式
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell_alignment = Alignment(vertical='center', wrap_text=True)

        fields = result.get('field_names', [])

        # 表头
        for col_idx, fld in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=fld)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据
        for row_idx, rec in enumerate(result.get('records', []), 2):
            for col_idx, fld in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(fld, ''))
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 冻结首行 + 自动筛选 + 列宽
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for col_idx, fld in enumerate(fields, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(len(fld) * 2, 10), 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = os.path.splitext(os.path.basename(filepath))[0] + '.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """从 Excel 导入"""
    import openpyxl

    filepath = request.form.get('file', '')
    target = request.form.get('target', 'new')

    if 'excel_file' not in request.files:
        return jsonify({'error': '未上传文件'})

    excel_file = request.files['excel_file']

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel 文件为空'})

        headers = [str(h).strip() if h else f'COL_{i}' for i, h in enumerate(rows[0])]
        data_rows = [r for r in rows[1:] if any(r)]

        if target == 'new':
            # 创建新 DBF 文件
            new_name = f'import_{int(time.time())}.dbf'
            new_path = UPLOAD_DIR / new_name
            fields_info = [{'name': h, 'type': 'C', 'length': 200, 'decimal': 0} for h in headers if h]
            result = create_dbf_file(new_path, fields_info)
            if 'error' in result:
                return jsonify(result)

            # 写入数据
            table = _safe_open_write(new_path)
            for row in data_rows:
                row_data = [str(v) if v is not None else '' for v in row]
                while len(row_data) < len(table.field_names):
                    row_data.append('')
                row_data = row_data[:len(table.field_names)]
                table.append(tuple(row_data))
            table.close()

            info = get_dbf_info(new_path)
            return jsonify({'success': True, 'path': str(new_path), 'info': info})

        else:
            # 追加到现有 DBF
            if not filepath or not os.path.exists(filepath):
                return jsonify({'error': '目标 DBF 文件不存在'})

            table = _safe_open_read(filepath)
            dbf_headers = list(table.field_names)
            table.close()

            # 列映射
            col_map = {}
            for i, h in enumerate(headers):
                if h in dbf_headers:
                    col_map[i] = h

            if not col_map:
                return jsonify({'error': 'Excel 列名与 DBF 字段无匹配'})

            table = _safe_open_write(filepath)
            for row in data_rows:
                row_vals = [''] * len(table.field_names)
                for col_idx, fld_name in col_map.items():
                    if col_idx < len(row):
                        row_vals[table.field_names.index(fld_name)] = str(row[col_idx]) if row[col_idx] is not None else ''
                table.append(tuple(row_vals))
            table.close()
            return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/create', methods=['POST'])
def create_table():
    """创建新的 DBF 文件"""
    data = request.json
    filename = data.get('filename', 'new.dbf')
    fields = data.get('fields', [])

    if not filename.endswith('.dbf'):
        filename += '.dbf'

    new_path = UPLOAD_DIR / filename
    result = create_dbf_file(new_path, fields)
    if 'error' not in result:
        result['path'] = str(new_path)
    return jsonify(result)


@app.route('/api/file/delete', methods=['POST'])
def delete_file():
    """删除 DBF 文件（仅允许删除 .dbf 文件，防止泄密）"""
    data = request.json
    filepath = data.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})
    
    # 安全检查：仅允许删除 .dbf 文件
    if not filepath.lower().endswith('.dbf'):
        return jsonify({'error': '仅允许删除 .dbf 文件'})
    
    try:
        os.chmod(filepath, 0o666)
        os.remove(filepath)
        # 同时删除可能的缓存文件
        for ext in ['.tmp', '.bak']:
            cache_path = filepath + ext
            if os.path.exists(cache_path):
                os.chmod(cache_path, 0o666)
                os.remove(cache_path)
        return jsonify({'success': True, 'deleted': os.path.basename(filepath)})
    except Exception as e:
        return jsonify({'error': f'删除失败: {e}'})


@app.route('/api/export/dbf')
def export_dbf():
    """导出 DBF 文件（保持原格式，直接下载）"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})
    basename = os.path.basename(filepath)
    return send_file(
        filepath,
        mimetype='application/octet-stream',
        as_attachment=True,
        download_name=basename
    )


@app.route('/api/download')
def download_dbf():
    """下载 DBF 文件"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'})
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


@app.route('/api/upload', methods=['POST'])
def upload_dbf():
    """上传 DBF 文件"""
    if 'dbf_file' not in request.files:
        return jsonify({'error': '未上传文件'})
    f = request.files['dbf_file']
    if not f.filename.lower().endswith('.dbf'):
        return jsonify({'error': '只支持 .dbf 文件'})

    save_path = UPLOAD_DIR / f.filename
    f.save(str(save_path))
    info = get_dbf_info(save_path)
    return jsonify({'success': True, 'path': str(save_path), 'info': info})


@app.route('/')
def index():
    """主页"""
    html_path = TEMPLATES_DIR / 'index.html'
    if html_path.exists():
        return html_path.read_text(encoding='utf-8')
    return '<h1>DBF Editor - templates/index.html not found</h1>'


if __name__ == '__main__':
    import socket

    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return '127.0.0.1'

    local_ip = get_local_ip()

    print()
    print('=' * 50)
    print('  DBF 编辑器 v2.0')
    print('=' * 50)
    print(f'  本地访问:   http://127.0.0.1:{PORT}')
    print(f'  局域网访问: http://{local_ip}:{PORT}')
    print('=' * 50)
    print('  按 Ctrl+C 停止服务器')
    print()

    app.run(host='0.0.0.0', port=PORT, debug=False)
